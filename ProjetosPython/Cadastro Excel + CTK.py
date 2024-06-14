import customtkinter as s
import datetime
data = datetime.date.today()
data= data.strftime("%d/%m/%Y")
import pandas as pd
from tkinter import ttk
import openpyxl.workbook
import openpyxl
from openpyxl import Workbook,load_workbook

#Adicionar:
#Versionamento

def EnvioExcel(nome,servicos,dataVenda,pago,formaPagamento,telefone,primeiroCadastro): #Definindo a função
    statusvar.set("") #Limpa a variavel de status
    try: #Tenta acessar a tabela
        book = load_workbook('data.xlsx') #Define o book ultilizado
        sheet = book['Sheet'] #Define a pagina utilizada
        statusvar.set("Tabela encontrada")
    except:
        clientes = ["nome", "servicos", "dataVenda", "pago", "formaPagamento", "telefone", "primeiroCadastro"]
        book = openpyxl.Workbook()
        sheet = book['Sheet']
        sheet.append(clientes) #Define o nome das colunas
        book.save('data.xlsx') 
        statusvar.set("Nova tabela criada")
    
    sheet.column_dimensions['A'].width = 25 #Seta o tamanho horizontal das colunas
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 11
    sheet.column_dimensions['F'].width = 12

    df = pd.read_excel('data.xlsx',sheet_name = 'Sheet') #guarda o arquivo excel em um dataframe
    if nome in df["nome"].values and dataVenda == data: #Checa se o nome ja foi cadastrado no dia
        statusvar.set("Cadastro repetido")

    elif nome == "" or pago == "":
        statusvar.set("Falta algo!")
        
    else:
        try:
            sheet.append([nome,servicos,dataVenda,pago,formaPagamento,telefone,primeiroCadastro]) #Guarda as opções passadas na função em uma linha do excel
            book.save('data.xlsx')
            statusvar.set("Cadastrado com sucesso!")
        except:
            statusvar.set("Feche o Excel!")

def Voltar(j):
    janelaMenu.deiconify()
    j.withdraw()

def funCadastros():
    janelaMenu.withdraw()
    JanelaCadastro = s.CTkToplevel(janelaMenu)
    JanelaCadastro.geometry("600x300") #Definindo o tamanho inicial da janela
    JanelaCadastro.title("Clientes Cadastrados") #Definindo nome da janela
    JanelaCadastro.resizable(False,False)
    s.set_appearance_mode("Dark")  
    s.set_default_color_theme("dark-blue")
    BTNVoltar = s.CTkButton(JanelaCadastro,text="Voltar",command=lambda:Voltar(JanelaCadastro))
    BTNVoltar.grid(row=1,column=1)


janelaMenu = s.CTk() #Criando a janela
janelaMenu.geometry("630x250") #Definindo o tamanho inicial da janela
janelaMenu.title("Cadastro de Serviços") #Definindo nome da janela
janelaMenu.resizable(False,False) #Trava a janela no tamanho definido
s.set_appearance_mode("Dark")  #Modo de aparencia do programa (Light ou Dark)
s.set_default_color_theme("dark-blue") #Modo de esquema de cores

statusvar = s.StringVar()
statusvar.set("Conectado!")
LabelCtrl = s.CTkLabel(janelaMenu, textvariable=statusvar, padx=10)
LabelCtrl.grid(row=3, column=4)

LabelNome  = s.CTkLabel(janelaMenu, text="Nome:") #Label "Nome"
LabelNome.grid(row=1,column=1,pady=10, padx=14)

EntryNome = s.CTkEntry(janelaMenu, placeholder_text="Digite o nome", width=250) #caixa de entrada de nome
EntryNome.grid(row=1, column=2,pady=10,sticky="W")

LabelValor = s.CTkLabel(janelaMenu,text="Valor:") #Label "Valor"
LabelValor.grid(row=1,column=3, padx=10)

EntryValor = s.CTkEntry(janelaMenu, placeholder_text="Valor...", width=100) #Caixa de entrada do valor
EntryValor.grid(row=1, column=4,pady=10,sticky="W")

LabelNumero = s.CTkLabel(janelaMenu, text="Celular:") #Label "Celular"
LabelNumero.grid(row=2,column=1)

EntryNumero = s.CTkEntry(janelaMenu, placeholder_text="(99)99999-9999") #Caixa de entrada de numeros
EntryNumero.grid(row=2, column=2, sticky="W", padx=5,pady=10)

OPTPagamento = s.CTkOptionMenu(janelaMenu, values=["Outro","Cartão","Dinheiro","Pix"],width= 85)  #Menu de opções de pagamento
OPTPagamento.grid(row=1,column=5, padx=10)

LabelOBS = s.CTkLabel(janelaMenu, text="Serviços:") #Label "Serviços"
LabelOBS.grid(row=3, column=1)

LabelData = s.CTkLabel(janelaMenu,text=data) 
LabelData.grid(row=6,column=1,padx=10, sticky="S")

EntryOBS = s.CTkTextbox(janelaMenu,height=100,width=300) #Caixa de entrada para serviços
EntryOBS.grid(row=3,column=2,columnspan=2,rowspan=3,pady=10)

ChkVisita = s.CTkCheckBox(janelaMenu,text="Primeiro Cadastro") #Check box primeiro cadastro
ChkVisita.grid(row=2, column=4)
    
BTNSalvar = s.CTkButton(janelaMenu,text="Salvar",width=100,command=lambda:EnvioExcel(EntryNome.get(),EntryOBS.get("1.0", "end-1c"),data,EntryValor.get(),OPTPagamento.get(),EntryNumero.get(), ChkVisita.get())) #Botão salvar
BTNSalvar.grid(row=6,column=5)

BTNVer = s.CTkButton(janelaMenu,text="Ver cadastros",command=funCadastros,width=100) #Botao "Ver cadastros"
BTNVer.grid(row=6,column=4,padx=5)

janelaMenu.mainloop()